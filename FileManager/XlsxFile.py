import openpyxl as xls
import numpy as np


class XlsFile:
    def __init__(self, path):
        self._xls_file = path

    def save_position(self,x, y, angle, timestamp, linear_velocity, angular_velocity, x_error, y_error, angle_error):
        xls_file = xls.Workbook()
        sheet = xls_file.active
        sheet.cell(row=2, column=1).value = 'MULTIROBOT BASE STATION CONTROL ENVIRONMENT'
        sheet.cell(row=4, column=1).value = 'ROBOT ID'
        sheet.cell(row=4, column=3).value = 'DATE'
        sheet.cell(row=6, column=1).value = 'ROBOT POSE'
        sheet.cell(row=7, column=1).value = 'Timestamp'
        sheet.cell(row=7, column=2).value = 'Position X'
        sheet.cell(row=7, column=3).value = 'Position Y'
        sheet.cell(row=7, column=4).value = 'Angle Phi'
        sheet.cell(row=7, column=5).value = 'Linear velocity'
        sheet.cell(row=7, column=6).value = 'Angular velocity'
        sheet.cell(row=7, column=7).value = 'X error'
        sheet.cell(row=7, column=8).value = 'Y error'
        sheet.cell(row=7, column=9).value = 'Angle error'
        for data in range(len(x)):
            sheet.cell(row=data + 8, column=1).value = timestamp[data]
            sheet.cell(row=data + 8, column=2).value = x[data]
            sheet.cell(row=data + 8, column=3).value = y[data]
            sheet.cell(row=data + 8, column=4).value = angle[data]
            sheet.cell(row=data + 8, column=5).value = linear_velocity[data]
            sheet.cell(row=data + 8, column=6).value = angular_velocity[data]
            sheet.cell(row=data + 8, column=7).value = x_error[data]
            sheet.cell(row=data + 8, column=8).value = y_error[data]
            sheet.cell(row=data + 8, column=9).value = angle_error[data]
        xls_file.save(self._xls_file)

    def read_position(self):
        file = xls.load_workbook(self._xls_file)
        self._names_sheets = file.get_sheet_names()
        self._sheet = file.get_sheet_by_name(self._names_sheets[0])
        row = 8
        pose_components = 4
        x = []
        y = []
        angle = []
        timestamp = []
        while True:
            if not self._sheet.cell(row=row,column=1).value:
                break
            pose_temporary = np.zeros(pose_components)
            for component in range(pose_components):
                pose_temporary[component] = self._sheet.cell(row=row,column=component+1).value
            timestamp.append(pose_temporary[0])
            x.append(pose_temporary[1])
            y.append(pose_temporary[2])
            angle.append(pose_temporary[3])
            row += 1
        file.close()
        return x, y, angle, timestamp
